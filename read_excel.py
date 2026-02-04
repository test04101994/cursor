"""
Functions to read Excel files cell-wise and check for blank/null/N/A values.
Supports both local file paths and AWS S3 URIs.
"""

from typing import Any, Optional, List, Dict
from openpyxl import load_workbook
from utils import get_local_path, is_blank_or_na, cleanup_temp_file


def read_cell(excel_path: str, cell_address: str, sheet_name: Optional[str] = None) -> Any:
    """
    Read a specific cell from an Excel file.
    
    Args:
        excel_path: Path to the Excel file (local path or S3 URI like 's3://bucket/file.xlsx')
        cell_address: Cell address (e.g., 'A1', 'B5', 'C10')
        sheet_name: Optional sheet name. If None, uses the active sheet.
    
    Returns:
        The cell value, or None if cell is empty
    
    Raises:
        FileNotFoundError: If file not found
        ValueError: If sheet name not found
    
    Example:
        # Local file
        value = read_cell('sample_data.xlsx', 'C2')
        value = read_cell('sample_data.xlsx', 'B5', sheet_name='Sheet1')
        
        # S3 URI
        value = read_cell('s3://my-bucket/data/file.xlsx', 'C2')
    """
    local_path, is_temp = get_local_path(excel_path)
    
    try:
        workbook = load_workbook(local_path, data_only=True)
        
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                available = workbook.sheetnames
                workbook.close()
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {available}"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        cell = sheet[cell_address]
        value = cell.value
        
        workbook.close()
        return value
    finally:
        if is_temp:
            cleanup_temp_file(local_path)


def read_cell_range(excel_path: str, start_cell: str, end_cell: str,
                   sheet_name: Optional[str] = None) -> List[Any]:
    """
    Read a range of cells from an Excel file.
    
    Args:
        excel_path: Path to the Excel file (local path or S3 URI)
        start_cell: Starting cell address (e.g., 'A1')
        end_cell: Ending cell address (e.g., 'C10')
        sheet_name: Optional sheet name. If None, uses the active sheet.
    
    Returns:
        List of cell values in row-major order
    
    Raises:
        FileNotFoundError: If file not found
        ValueError: If sheet name not found
    
    Example:
        # Local file
        values = read_cell_range('sample_data.xlsx', 'B2', 'C5')
        
        # S3 URI
        values = read_cell_range('s3://my-bucket/data/file.xlsx', 'B2', 'C5')
    """
    local_path, is_temp = get_local_path(excel_path)
    
    try:
        workbook = load_workbook(local_path, data_only=True)
        
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                available = workbook.sheetnames
                workbook.close()
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {available}"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        values = []
        for row in sheet[start_cell:end_cell]:
            for cell in row:
                values.append(cell.value)
        
        workbook.close()
        return values
    finally:
        if is_temp:
            cleanup_temp_file(local_path)


def is_cell_blank(excel_path: str, cell_address: str,
                  sheet_name: Optional[str] = None) -> bool:
    """
    Check if a specific cell in Excel is blank, null, or N/A.
    Returns only True or False.
    
    Args:
        excel_path: Path to the Excel file (local path or S3 URI)
        cell_address: Cell address (e.g., 'A1', 'B5', 'C10')
        sheet_name: Optional sheet name. If None, uses the active sheet.
    
    Returns:
        True if cell is blank/null/N/A, False otherwise
    
    Raises:
        FileNotFoundError: If file not found
        ValueError: If sheet name not found
    
    Example:
        # Local file
        if is_cell_blank('sample_data.xlsx', 'C5'):
            # Handle blank cell
        
        # S3 URI
        if is_cell_blank('s3://my-bucket/data/file.xlsx', 'C5'):
            # Handle blank cell
    """
    value = read_cell(excel_path, cell_address, sheet_name)
    return is_blank_or_na(value)


def check_cell_value(excel_path: str, cell_address: str,
                    sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    Read a cell and check if it's blank, null, or N/A.
    
    Args:
        excel_path: Path to the Excel file (local path or S3 URI)
        cell_address: Cell address (e.g., 'A1', 'B5')
        sheet_name: Optional sheet name. If None, uses the active sheet.
    
    Returns:
        Dictionary with:
        - 'value': The cell value
        - 'is_blank': True if blank/null/N/A
        - 'cell_address': The cell address
        - 'data_type': Type of the value
    
    Raises:
        FileNotFoundError: If file not found
        ValueError: If sheet name not found
    
    Example:
        # Local file
        result = check_cell_value('sample_data.xlsx', 'C5')
        if result['is_blank']:
            # Handle blank cell
        
        # S3 URI
        result = check_cell_value('s3://my-bucket/data/file.xlsx', 'C5')
    """
    value = read_cell(excel_path, cell_address, sheet_name)
    is_blank = is_blank_or_na(value)
    
    return {
        'value': value,
        'is_blank': is_blank,
        'cell_address': cell_address,
        'data_type': type(value).__name__ if value is not None else 'NoneType'
    }


def read_all_cells_in_column(excel_path: str, column: str,
                             start_row: int = 1, end_row: Optional[int] = None,
                             sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Read all cells in a specific column.
    
    Args:
        excel_path: Path to the Excel file (local path or S3 URI)
        column: Column letter (e.g., 'B', 'C')
        start_row: Starting row number (default: 1)
        end_row: Ending row number. If None, reads until first empty cell.
        sheet_name: Optional sheet name. If None, uses the active sheet.
    
    Returns:
        List of dictionaries with cell info:
        [{'address': 'B1', 'value': ..., 'is_blank': ..., 'row': ...}, ...]
    
    Raises:
        FileNotFoundError: If file not found
        ValueError: If sheet name not found
    
    Example:
        # Local file
        cells = read_all_cells_in_column('sample_data.xlsx', 'C', start_row=2)
        
        # S3 URI
        cells = read_all_cells_in_column('s3://my-bucket/data/file.xlsx', 'C', start_row=2)
    """
    local_path, is_temp = get_local_path(excel_path)
    max_rows = 10000
    
    try:
        workbook = load_workbook(local_path, data_only=True)
        
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                available = workbook.sheetnames
                workbook.close()
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {available}"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        results = []
        row = start_row
        
        while True:
            if end_row and row > end_row:
                break
            
            cell_address = f"{column}{row}"
            cell = sheet[cell_address]
            value = cell.value
            
            if value is None and row > start_row:
                if row > start_row + 1:
                    break
            
            results.append({
                'address': cell_address,
                'value': value,
                'is_blank': is_blank_or_na(value),
                'row': row
            })
            
            row += 1
            if row > max_rows:
                break
        
        workbook.close()
        return results
    finally:
        if is_temp:
            cleanup_temp_file(local_path)


def _column_index_to_letter(col_idx: int) -> str:
    """
    Convert 0-based column index to Excel column letter(s).
    
    Args:
        col_idx: 0-based column index (0 = A, 1 = B, ..., 25 = Z, 26 = AA, etc.)
    
    Returns:
        Excel column letter(s) (e.g., 'A', 'B', 'AA', 'AB')
    """
    result = ""
    col_idx += 1  # Convert to 1-based for calculation
    
    while col_idx > 0:
        col_idx -= 1
        result = chr(ord('A') + (col_idx % 26)) + result
        col_idx //= 26
    
    return result


def _letter_to_column_index(column: str) -> int:
    """
    Convert Excel column letter(s) to 0-based column index.
    
    Args:
        column: Excel column letter(s) (e.g., 'A', 'B', 'AA', 'AB')
    
    Returns:
        0-based column index
    """
    col_num = 0
    for char in column.upper():
        col_num = col_num * 26 + (ord(char) - ord('A') + 1)
    return col_num - 1  # Make it 0-based


def read_all_cells_in_row(excel_path: str, row: int,
                          start_column: str = 'A',
                          end_column: Optional[str] = None,
                          sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Read all cells in a specific row.
    
    Args:
        excel_path: Path to the Excel file (local path or S3 URI)
        row: Row number (1-based, Excel notation)
        start_column: Starting column letter (default: 'A')
        end_column: Ending column letter. If None, reads until first empty cell.
        sheet_name: Optional sheet name. If None, uses the active sheet.
    
    Returns:
        List of dictionaries with cell info:
        [{'address': 'A2', 'value': ..., 'is_blank': ..., 'column': ...}, ...]
    
    Raises:
        FileNotFoundError: If file not found
        ValueError: If sheet name not found
    
    Example:
        # Local file - read row 2 from column B until empty
        cells = read_all_cells_in_row('sample_data.xlsx', 2, start_column='B')
        
        # Read row 5 from column A to column D
        cells = read_all_cells_in_row('sample_data.xlsx', 5, start_column='A', end_column='D')
        
        # S3 URI
        cells = read_all_cells_in_row('s3://my-bucket/data/file.xlsx', 2, start_column='B')
    """
    local_path, is_temp = get_local_path(excel_path)
    max_cols = 10000
    
    try:
        workbook = load_workbook(local_path, data_only=True)
        
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                available = workbook.sheetnames
                workbook.close()
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {available}"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        start_col_idx = _letter_to_column_index(start_column)
        end_col_idx = None
        if end_column:
            end_col_idx = _letter_to_column_index(end_column)
        
        results = []
        col_idx = start_col_idx
        
        while True:
            if end_col_idx is not None and col_idx > end_col_idx:
                break
            
            col_letter = _column_index_to_letter(col_idx)
            cell_address = f"{col_letter}{row}"
            
            try:
                cell = sheet[cell_address]
                value = cell.value
            except (AttributeError, IndexError):
                value = None
            
            # Stop if we hit empty cell and we're past start_column
            if value is None and col_idx > start_col_idx:
                if end_col_idx is None:  # Only stop early if no end_column specified
                    break
            
            results.append({
                'address': cell_address,
                'value': value,
                'is_blank': is_blank_or_na(value),
                'column': col_letter
            })
            
            col_idx += 1
            if col_idx > max_cols:
                break
        
        workbook.close()
        return results
    finally:
        if is_temp:
            cleanup_temp_file(local_path)
