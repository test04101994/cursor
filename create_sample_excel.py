#!/usr/bin/env python3
"""
Create a sample Excel file for testing.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_sample_excel(filename='sample_data.xlsx'):
    """Create a sample Excel file with various data types and empty cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Data"
    
    # Header row
    ws['A1'] = 'Row'
    ws['B1'] = 'Key'
    ws['C1'] = 'Value'
    ws['D1'] = 'Status'
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
    
    # Sample data with various scenarios
    data = [
        ('name', 'John Doe', 'Valid'),
        ('email', 'john@example.com', 'Valid'),
        ('age', 30, 'Valid'),
        ('city', 'New York', 'Valid'),
        ('phone', None, 'Empty'),  # None value
        ('address', '', 'Empty'),  # Empty string
        ('status', 'N/A', 'N/A'),  # N/A value
        ('department', 'Engineering', 'Valid'),
        ('salary', 75000, 'Valid'),
        ('notes', None, 'Empty'),  # None value
        ('active', True, 'Valid'),
        ('tags', '["Python", "JavaScript"]', 'Valid'),
    ]
    
    # Write data starting from row 2
    for idx, (key, value, status) in enumerate(data, start=2):
        ws[f'A{idx}'] = idx - 1
        ws[f'B{idx}'] = key
        ws[f'C{idx}'] = value
        ws[f'D{idx}'] = status
    
    # Add some empty rows
    ws['B15'] = 'empty_key'
    ws['C15'] = None
    
    ws['B16'] = 'na_value'
    ws['C16'] = 'N/A'
    
    ws['B17'] = 'null_value'
    ws['C17'] = 'null'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    
    wb.save(filename)
    return filename

if __name__ == '__main__':
    create_sample_excel()

